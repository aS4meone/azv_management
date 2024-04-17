import requests
from openpyxl.styles import Border, Side
from openpyxl.workbook import Workbook


def fetch_data(url):
    response = requests.get(url)
    if response.status_code == 200:
        return response.json()
    else:
        print("Failed to fetch data from the endpoint.")
        return None


def write_to_excel(data):
    wb = Workbook()
    ws = wb.active
    ws.append(["Наименование", "Цена"])

    max_name_length = 0  # Для хранения максимальной длины значения в столбце "Наименование"

    for item in data:
        name = item["name"]
        price = item["price"]
        if item["quantity"] < 10:
            name += f" ({item['quantity']})"

        # Определение максимальной длины
        max_name_length = max(max_name_length, len(name))

        ws.append([name, price])

    # Установка ширины столбца "Наименование"
    ws.column_dimensions['A'].width = max_name_length

    # Добавляем обводку таблицы
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

    wb.save("price.xlsx")
    print("Data has been written to price.xlsx")


def write_to_excel_default(data):
    wb = Workbook()
    ws = wb.active
    ws.append(["Наименование", "Количество", "Цена"])

    max_name_length = 0  # Для хранения максимальной длины значения в столбце "Наименование"

    for item in data:
        name = item["name"]
        quantity = item["quantity"]
        price = item["price"]
        if quantity < 10:
            name += f" ({quantity})"

        max_name_length = max(max_name_length, len(name))

        ws.append([name, quantity, price])

    ws.column_dimensions['A'].width = max_name_length

    # Добавляем обводку таблицы
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

    wb.save("price_default.xlsx")
    print("Data has been written to price_default.xlsx")
